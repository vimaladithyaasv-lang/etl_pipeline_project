# ============================================================
#  EXCEL DATA DASHBOARD  —  Full Analyst Edition v2
#  Features: Smart parsing, Data Quality, Stats, Outliers,
#            Group-by, Chart Builder, Time Series, AI Summary,
#            Custom Query + GROUP BY, Pivot Table, PDF/Excel
#            Export, MySQL Load/Save, Upload History, PII Masking
#            FastAPI Import: saves ALL sheets to DB on button click
#
#  CHANGES v2.1:
#  • "Import All Sheets" button → cleans + saves ALL sheets to SQLite
#    via FastAPI (same logic as the old "Clean & Save" for every sheet)
#  • Tab 8 "Clean & Save to Database" button REMOVED →
#    replaced with "🧹 Clean Data" (preview only, no DB write)
# ============================================================

import pandas as pd
from sqlalchemy import create_engine, text, inspect
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import numpy as np
import re
import io
import json
import requests as http_requests
from scipy import stats as scipy_stats
try:
    import pandasql as psql
    SQL_SUPPORTED = True
except ImportError:
    SQL_SUPPORTED = False

# ── Page config ────────────────────────────────────────────
st.set_page_config(page_title="Excel Analyst Dashboard", page_icon="📊", layout="wide")

# ── CSS ────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;}
#MainMenu,footer,header{visibility:hidden;}
.stApp{background:#f7f8fc;}

.dashboard-header{
    background:linear-gradient(135deg,#1a1f36 0%,#2d3561 100%);
    border-radius:16px;padding:28px 36px;margin-bottom:24px;
    display:flex;align-items:center;gap:16px;
    box-shadow:0 8px 32px rgba(26,31,54,.18);
}
.dashboard-header h1{color:#fff;font-size:1.9rem;font-weight:700;margin:0;letter-spacing:-.5px;}
.dashboard-header p{color:rgba(255,255,255,.6);font-size:.9rem;margin:4px 0 0 0;}
.header-icon{font-size:2.4rem;}

.section-header{
    display:flex;align-items:center;gap:10px;
    font-size:1.05rem;font-weight:700;color:#1a1f36;
    margin:22px 0 12px 0;padding-bottom:8px;
    border-bottom:2px solid #e8eaf0;
}
.section-icon{font-size:1.1rem;}

.metric-card{
    background:#fff;border-radius:14px;padding:18px 20px;
    box-shadow:0 2px 12px rgba(26,31,54,.07);
    border:1px solid #eceef5;transition:box-shadow .2s;height:100%;
}
.metric-card:hover{box-shadow:0 6px 24px rgba(26,31,54,.13);}
.metric-label{font-size:.75rem;font-weight:600;color:#8a8fa8;text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px;}
.metric-value{font-size:1.85rem;font-weight:700;color:#1a1f36;font-family:'DM Mono',monospace;line-height:1.1;}
.metric-sub{font-size:.8rem;margin-top:4px;font-weight:500;}
.sub-green{color:#22c55e;} .sub-red{color:#ef4444;} .sub-blue{color:#6366f1;} .sub-gray{color:#8a8fa8;}
.metric-detail{font-size:.73rem;color:#8a8fa8;margin-top:2px;}

.import-btn-wrap{background:linear-gradient(135deg,#0f172a,#1e3a5f);border-radius:14px;padding:20px 24px;margin:12px 0;display:flex;align-items:center;gap:16px;}
.import-btn-wrap p{color:rgba(255,255,255,.75);font-size:.85rem;margin:4px 0 0 0;}
.import-btn-wrap h3{color:#fff;margin:0;font-size:1rem;}

.stButton>button{
    background:linear-gradient(135deg,#4f46e5,#7c3aed);color:#fff;
    border:none;border-radius:10px;padding:10px 22px;
    font-size:.9rem;font-weight:600;font-family:'DM Sans',sans-serif;
    cursor:pointer;transition:all .2s;
    box-shadow:0 4px 14px rgba(79,70,229,.35);width:100%;
}
.stButton>button:hover{transform:translateY(-1px);box-shadow:0 6px 20px rgba(79,70,229,.45);}

.import-result-row{background:#fff;border:1px solid #e0e7ff;border-radius:10px;padding:10px 16px;margin:4px 0;display:flex;justify-content:space-between;align-items:center;}
.import-result-row .sheet-name{font-weight:600;color:#3730a3;font-size:.9rem;}
.import-result-row .row-count{color:#6b7280;font-size:.82rem;}
.import-result-row .status-ok{color:#16a34a;font-weight:700;}
.import-result-row .status-err{color:#dc2626;font-weight:700;}

.badge{display:inline-block;background:#dcfce7;color:#15803d;border-radius:6px;padding:3px 10px;font-size:.78rem;font-weight:600;}
.badge-red{background:#fee2e2;color:#991b1b;}
.badge-blue{background:#dbeafe;color:#1e40af;}
.badge-yellow{background:#fef9c3;color:#854d0e;}

.divider{border:none;border-top:1.5px solid #eceef5;margin:18px 0;}

.stTabs [data-baseweb="tab-list"]{gap:4px;background:#eef0fb;border-radius:10px;padding:4px;}
.stTabs [data-baseweb="tab"]{border-radius:7px;font-weight:600;color:#6366f1;padding:6px 16px;}
.stTabs [aria-selected="true"]{background:#fff !important;color:#1a1f36 !important;box-shadow:0 2px 8px rgba(26,31,54,.08);}

.custom-success{background:#f0fdf4;border:1.5px solid #86efac;border-radius:10px;padding:12px 16px;color:#15803d;font-weight:500;font-size:.88rem;margin:8px 0;}
.custom-info{background:#eff6ff;border:1.5px solid #93c5fd;border-radius:10px;padding:12px 16px;color:#1d4ed8;font-weight:500;font-size:.88rem;margin:8px 0;}
.custom-warning{background:#fffbeb;border:1.5px solid #fcd34d;border-radius:10px;padding:12px 16px;color:#92400e;font-weight:500;font-size:.88rem;margin:8px 0;}
.custom-error{background:#fef2f2;border:1.5px solid #fca5a5;border-radius:10px;padding:12px 16px;color:#991b1b;font-weight:500;font-size:.88rem;margin:8px 0;}

.outlier-badge{display:inline-block;background:#fee2e2;color:#991b1b;border-radius:6px;padding:2px 8px;font-size:.75rem;font-weight:700;margin-left:6px;}
.pii-badge{display:inline-block;background:#fef9c3;color:#854d0e;border-radius:6px;padding:2px 8px;font-size:.75rem;font-weight:700;margin-left:6px;}

.ai-box{background:linear-gradient(135deg,#f0f4ff,#faf5ff);border:1.5px solid #c7d2fe;border-radius:14px;padding:18px 22px;margin:10px 0;}
.ai-box h4{color:#4338ca;margin:0 0 8px 0;font-size:.95rem;}
.ai-box p{color:#374151;font-size:.88rem;line-height:1.6;margin:0;}

.history-row{background:#fff;border:1px solid #eceef5;border-radius:10px;padding:12px 16px;margin:6px 0;display:flex;justify-content:space-between;align-items:center;}
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════
#  CONFIGURATION
# ════════════════════════════════════════════════════════════
FASTAPI_BASE_URL = "http://127.0.0.1:8000"
SQLITE_DB_PATH   = "sqlite:///excel_dashboard.db"

# ════════════════════════════════════════════════════════════
#  BACKEND FUNCTIONS
# ════════════════════════════════════════════════════════════

# ── MySQL (port 3307) ────────────────────────────────
MYSQL_USER = "root"
MYSQL_PASSWORD = "Vimal@2730"  # ← change this
MYSQL_HOST = "127.0.0.1"
MYSQL_PORT = 3307
MYSQL_DB = "etl_project_db"  # ← change this to your database name

@st.cache_resource
def get_mysql_engine():
    return create_engine(
        "mysql+pymysql://",
        connect_args={
            "host": MYSQL_HOST,
            "port": MYSQL_PORT,
            "user": MYSQL_USER,
            "password": MYSQL_PASSWORD,
            "database": MYSQL_DB,
        }
    )

def get_saved_tables():
    try:
        engine = get_mysql_engine()
        insp = inspect(engine)
        return [t for t in insp.get_table_names() if not t.startswith("_")]
    except Exception:
        return []

def load_table_from_db(table_name):
    engine = get_mysql_engine()
    with engine.connect() as conn:
        return pd.read_sql(f'SELECT * FROM `{table_name}`', conn)

def save_df_to_mysql(df: pd.DataFrame, table_name: str) -> str:
    """Save a single DataFrame to MySQL directly."""
    engine = get_mysql_engine()
    safe_name = re.sub(r"[^\w]", "_", table_name)[:60]
    df.columns = [re.sub(r"[^\w]", "_", str(c)).strip("_") for c in df.columns]
    df.to_sql(safe_name, engine, if_exists="replace", index=False, method="multi")
    return safe_name

# ── Smart Header Detection ─────────────────────────────────
def detect_and_parse_sheet(xl, sheet_name, max_scan_rows=40):
    df_raw_full = xl.parse(sheet_name, header=None, dtype=str)
    best_row, best_score = 0, -1
    for i in range(min(max_scan_rows, len(df_raw_full))):
        row = df_raw_full.iloc[i]
        non_null = row.dropna().astype(str)
        non_null = non_null[~non_null.str.strip().str.lower().isin(["none", "nan", ""])]
        text_like = non_null[~non_null.str.match(r"^\s*-?\d+(\.\d+)?\s*$")]
        score = len(text_like)
        if score > best_score:
            best_score, best_row = score, i
    df = xl.parse(sheet_name, header=best_row)
    df = df.dropna(how="all").dropna(axis=1, how="all")
    mask = df.apply(lambda r: all(str(v).strip().lower() in ("none","nan","") for v in r), axis=1)
    df = df[~mask].reset_index(drop=True)
    df.columns = [str(c).strip() if not str(c).startswith("Unnamed") else f"Col_{i}"
                  for i, c in enumerate(df.columns)]
    for col in df.columns:
        original = df[col].copy()
        converted = pd.to_numeric(df[col], errors='coerce')
        non_null_orig = df[col].dropna()
        non_null_conv = converted.dropna()
        if len(non_null_orig) > 0:
            numeric_ratio = len(non_null_conv) / len(non_null_orig)
            if numeric_ratio >= 0.6:
                df[col] = converted
                continue
        try:
            dt_converted = pd.to_datetime(df[col], infer_datetime_format=True, errors='coerce')
            dt_ratio = dt_converted.notna().sum() / max(len(df[col].dropna()), 1)
            if dt_ratio >= 0.7:
                df[col] = dt_converted
                continue
        except Exception:
            pass
        df[col] = original
    return df, best_row

# ── Data Cleaning ──────────────────────────────────────────
def clean_data(df):
    orig = df.shape
    dup = df.duplicated().sum()
    miss_before = df.isnull().sum().sum()
    df = df.drop_duplicates().dropna(how="all")
    for col in df.columns:
        if df[col].dtype != object:
            df[col] = df[col].ffill().bfill()
        else:
            df[col] = df[col].fillna("Unknown")
    df.columns = df.columns.str.strip().str.lower().str.replace(r"\s+", "_", regex=True)
    for col in df.columns:
        if df[col].dtype == object:
            converted = pd.to_numeric(df[col], errors='coerce')
            ratio = converted.notna().sum() / max(df[col].notna().sum(), 1)
            if ratio >= 0.6:
                df[col] = converted
    miss_after = df.isnull().sum().sum()
    return df, {"Original Rows": orig[0], "Cleaned Rows": df.shape[0],
                "Columns": df.shape[1], "Duplicates Removed": int(dup),
                "Missing Before": int(miss_before), "Missing After": int(miss_after)}

# ── Metric Stats ───────────────────────────────────────────
def compute_metric_stats(series):
    s = series.dropna()
    if len(s) == 0:
        return {k: 0 for k in ["total","mean","median","min","max","std","cv","skew","kurt","count","non_null_pct"]}
    std = s.std()
    mean = s.mean()
    return {
        "total": s.sum(), "mean": mean, "median": s.median(),
        "min": s.min(), "max": s.max(), "std": std,
        "cv": (std/mean*100) if mean != 0 else 0,
        "skew": float(s.skew()), "kurt": float(s.kurtosis()),
        "count": len(s), "non_null_pct": len(s)/len(series)*100,
    }

# ── Data Quality Checks ────────────────────────────────────
def data_quality_report(df):
    issues = []
    num_cols = df.select_dtypes(include="number").columns.tolist()
    cat_cols = df.select_dtypes(include="object").columns.tolist()
    outlier_summary = {}
    for col in num_cols:
        s = df[col].dropna()
        Q1, Q3 = s.quantile(0.25), s.quantile(0.75)
        IQR = Q3 - Q1
        out = df[(df[col] < Q1 - 1.5*IQR) | (df[col] > Q3 + 1.5*IQR)]
        if len(out) > 0:
            outlier_summary[col] = {"count": len(out), "pct": len(out)/len(df)*100,
                                     "rows": out.index.tolist()[:10]}
            issues.append(f"⚠️ **{col}**: {len(out)} outliers ({len(out)/len(df)*100:.1f}%)")
    id_like = [c for c in df.columns if any(k in c.lower() for k in ["id","code","no","num","ref"])]
    key_dups = {}
    for col in id_like:
        d = df[col].duplicated().sum()
        if d > 0:
            key_dups[col] = d
            issues.append(f"🔑 **{col}**: {d} duplicate key values")
    string_issues = {}
    for col in cat_cols:
        variants = df[col].dropna().astype(str)
        lower_counts = variants.str.lower().value_counts()
        raw_counts = variants.value_counts()
        if len(lower_counts) < len(raw_counts):
            diff = len(raw_counts) - len(lower_counts)
            string_issues[col] = diff
            issues.append(f"🔤 **{col}**: {diff} case inconsistencies")
    card_warn = {}
    for col in num_cols:
        unique_vals = df[col].nunique()
        if unique_vals <= 5 and unique_vals > 0:
            card_warn[col] = unique_vals
            issues.append(f"🔢 **{col}**: only {unique_vals} unique values — may be a category")
    date_cols = df.select_dtypes(include=["datetime64"]).columns.tolist()
    freshness = {}
    for col in date_cols:
        latest = df[col].max()
        days_old = (pd.Timestamp.now() - latest).days if pd.notnull(latest) else None
        freshness[col] = {"latest": latest, "days_old": days_old}
        if days_old and days_old > 90:
            issues.append(f"📅 **{col}**: latest date is {latest.date()} ({days_old} days ago)")
    return {
        "issues": issues,
        "outliers": outlier_summary,
        "key_dups": key_dups,
        "string_issues": string_issues,
        "card_warn": card_warn,
        "freshness": freshness,
    }

# ── PII Detection ──────────────────────────────────────────
PII_KEYWORDS = ["email","phone","mobile","ssn","passport","dob","birth","address",
                "salary","account","credit","pan","aadhar","aadhaar","national_id",
                "gender","age","name","employee_id","emp_id"]

def detect_pii_columns(df):
    pii = []
    for col in df.columns:
        col_lower = col.lower().replace(" ","_")
        if any(kw in col_lower for kw in PII_KEYWORDS):
            pii.append(col)
    return pii

def mask_pii(df, pii_cols):
    df2 = df.copy()
    for col in pii_cols:
        if col in df2.columns:
            df2[col] = df2[col].astype(str).apply(
                lambda x: x[:2] + "*"*(max(len(x)-4,2)) + x[-2:] if len(x) > 4 else "****"
            )
    return df2

# ── Auto Charts ────────────────────────────────────────────
def auto_charts(df):
    num_cols = df.select_dtypes(include="number").columns.tolist()
    cat_cols = df.select_dtypes(include="object").columns.tolist()
    date_cols = df.select_dtypes(include=["datetime64"]).columns.tolist()
    charts = []
    layout = dict(plot_bgcolor="#fff", paper_bgcolor="#fff", font_family="DM Sans", title_font_size=14)
    for col in num_cols[:3]:
        fig = px.histogram(df, x=col, nbins=30, title=f"Distribution — {col}", color_discrete_sequence=["#6366f1"])
        fig.update_layout(**layout); charts.append((f"Histogram: {col}", fig))
    for col in num_cols[:3]:
        fig = px.box(df, y=col, title=f"Box Plot — {col}", color_discrete_sequence=["#7c3aed"])
        fig.update_layout(**layout); charts.append((f"Box Plot: {col}", fig))
    if cat_cols and num_cols:
        for cat in cat_cols[:2]:
            if df[cat].nunique() <= 20:
                grouped = df.groupby(cat, as_index=False)[num_cols[0]].mean().sort_values(by=num_cols[0], ascending=False).head(10)
                fig = px.bar(grouped, x=cat, y=num_cols[0], title=f"{cat} vs Avg {num_cols[0]}", color_discrete_sequence=["#0ea5e9"])
                fig.update_layout(**layout); charts.append((f"Bar: {cat} vs {num_cols[0]}", fig))
    if len(num_cols) >= 2:
        fig = px.scatter(df, x=num_cols[0], y=num_cols[1], title=f"Scatter: {num_cols[0]} vs {num_cols[1]}", color_discrete_sequence=["#f59e0b"])
        fig.update_layout(**layout); charts.append(("Scatter Plot", fig))
    for cat in cat_cols[:1]:
        if 2 <= df[cat].nunique() <= 8:
            vc = df[cat].value_counts().reset_index(); vc.columns = [cat, "count"]
            fig = px.pie(vc, names=cat, values="count", title=f"Pie — {cat}", color_discrete_sequence=px.colors.qualitative.Pastel)
            fig.update_layout(font_family="DM Sans", title_font_size=14); charts.append((f"Pie: {cat}", fig))
    if len(num_cols) >= 3:
        corr = df[num_cols].corr().round(2)
        fig = px.imshow(corr, text_auto=True, title="Correlation Heatmap", color_continuous_scale="RdBu_r")
        fig.update_layout(font_family="DM Sans", title_font_size=14); charts.append(("Correlation Heatmap", fig))
    if date_cols and num_cols:
        for dc in date_cols[:1]:
            fig = px.line(df.sort_values(dc), x=dc, y=num_cols[0], title=f"Time Series — {num_cols[0]}", color_discrete_sequence=["#6366f1"])
            fig.update_layout(**layout); charts.append((f"Time Series: {num_cols[0]}", fig))
    return charts

# ── Fuzzy Column Resolver ──────────────────────────────────
def resolve_col(df, name):
    def normalize(s):
        s = re.sub(r"([a-z])([A-Z])", r"\1_\2", str(s))
        s = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1_\2", s)
        return re.sub(r"[\s\-]+", "_", s).lower().strip()
    norm_name = normalize(name)
    for col in df.columns:
        if normalize(col) == norm_name:
            return col
    stripped = norm_name.replace("_", "")
    for col in df.columns:
        if normalize(col).replace("_", "") == stripped:
            return col
    for col in df.columns:
        nc = normalize(col)
        if norm_name in nc or nc in norm_name:
            return col
    return None

def resolve_cols(df, names):
    resolved, missing = [], []
    for n in names:
        r = resolve_col(df, n)
        if r:
            resolved.append(r)
        else:
            missing.append(n)
    return resolved, missing

# ── Custom Query Engine ────────────────────────────────────
def run_custom_query(df, query_str):
    query_str = query_str.strip()
    def col_hint():
        return f"Available columns: {list(df.columns)}"

    gb_pattern = re.compile(
        r"^(SUM|AVG|MEAN|COUNT|MIN|MAX|MEDIAN|STD)\s*\(([^\)]*)\)\s+GROUP\s+BY\s+([\w\s,]+?)(?:\s+WHERE\s+(.+))?$",
        re.IGNORECASE
    )
    m = gb_pattern.match(query_str)
    if m:
        func = m.group(1).upper()
        agg_col_raw = m.group(2).strip()
        group_raw = [c.strip() for c in m.group(3).split(",")]
        where_clause = m.group(4)
        sub_df = df
        if where_clause:
            try:
                sub_df = df.query(where_clause)
            except Exception as e:
                return None, f"❌ WHERE error: `{e}`"
        group_cols, missing_gc = resolve_cols(sub_df, group_raw)
        if missing_gc:
            return None, f"❌ Column(s) not found: {missing_gc}\n\n📋 {col_hint()}"
        agg_col = resolve_col(sub_df, agg_col_raw) if agg_col_raw not in ("", "*") else agg_col_raw
        if func != "COUNT" and not agg_col:
            return None, f"❌ Aggregation column `{agg_col_raw}` not found.\n\n📋 {col_hint()}"
        agg_map = {"SUM": "sum","AVG": "mean","MEAN": "mean","MIN": "min",
                   "MAX": "max","MEDIAN": "median","STD": "std","COUNT": "count"}
        try:
            if func == "COUNT":
                result_df = sub_df.groupby(group_cols).size().reset_index(name="COUNT")
            else:
                result_df = sub_df.groupby(group_cols)[agg_col].agg(agg_map[func]).reset_index()
                result_df.columns = group_cols + [f"{func}({agg_col})"]
            result_df = result_df.sort_values(result_df.columns[-1], ascending=False)
            return result_df, f"✅ GROUP BY returned **{len(result_df)}** groups."
        except Exception as e:
            return None, f"❌ GROUP BY error: `{e}`"

    agg_pattern = re.compile(
        r"^(SUM|AVG|MEAN|COUNT|MIN|MAX|MEDIAN|STD)\s*\(\s*([^\)]*)\s*\)\s*(?:WHERE\s+(.+))?$",
        re.IGNORECASE
    )
    m = agg_pattern.match(query_str)
    if m:
        func = m.group(1).upper()
        col_name_raw = m.group(2).strip()
        where_clause = m.group(3)
        sub_df = df
        if where_clause:
            try:
                sub_df = df.query(where_clause)
            except Exception as e:
                return None, f"❌ WHERE error: `{e}`"
        if func == "COUNT":
            result = len(sub_df) if col_name_raw in ("","*") else sub_df[col_name_raw].count()
            return sub_df, f"**COUNT = `{int(result):,}`**"
        col_name = resolve_col(df, col_name_raw)
        if not col_name:
            return None, f"❌ Column `{col_name_raw}` not found.\n\n📋 {col_hint()}"
        col_data = sub_df[col_name].dropna()
        ops = {"SUM": col_data.sum, "AVG": col_data.mean, "MEAN": col_data.mean,
               "MIN": col_data.min, "MAX": col_data.max, "MEDIAN": col_data.median, "STD": col_data.std}
        result = ops[func]()
        return sub_df, f"**{func}(`{col_name}`) = `{result:,.4f}`** ({len(col_data):,} rows)"

    top_m = re.match(r"^TOP\s+(\d+)\s+BY\s+(\w+)(?:\s+(ASC|DESC))?$", query_str, re.IGNORECASE)
    if top_m:
        n, col, order = int(top_m.group(1)), top_m.group(2), (top_m.group(3) or "DESC").upper()
        if col not in df.columns:
            return None, f"❌ Column `{col}` not found."
        result_df = df.nlargest(n, col) if order == "DESC" else df.nsmallest(n, col)
        return result_df, f"✅ Top {n} rows by `{col}` ({order})."

    try:
        result_df = df.query(query_str)
        return result_df, f"✅ Returned **{len(result_df):,}** of **{len(df):,}** rows."
    except Exception as e:
        return None, f"❌ Query error: `{e}`"

# ── Export to Excel ────────────────────────────────────────
def export_to_excel(df, sheet_name="Data"):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="1a1f36", end_color="1a1f36", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col_cells in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)
    buf.seek(0)
    return buf

# ── Upload History ─────────────────────────────────────────
def log_upload(filename, sheet, rows, cols):
    if "upload_history" not in st.session_state:
        st.session_state.upload_history = []
    st.session_state.upload_history.insert(0, {
        "file": filename, "sheet": sheet, "rows": rows,
        "cols": cols, "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })
    st.session_state.upload_history = st.session_state.upload_history[:20]

# ── FastAPI Import: call the /import-excel endpoint ───────
def call_fastapi_import(file_bytes: bytes, filename: str) -> dict:
    """
    POST the uploaded file to the FastAPI /import-excel endpoint.
    FastAPI will:
      1. Parse every sheet (smart header detection)
      2. Clean the data  (dedup, fill missing, normalise columns)
      3. Persist each sheet as a separate SQLite table (REPLACE if exists)
    Returns the JSON response from FastAPI.
    """
    try:
        resp = http_requests.post(
            f"{FASTAPI_BASE_URL}/import-excel",
            files={"file": (filename, file_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=120,
        )
        resp.raise_for_status()
        return resp.json()
    except http_requests.exceptions.ConnectionError:
        return {
            "status": "error",
            "message": (
                "FastAPI server is not running. "
                "Start it with: uvicorn api:app --reload"
            ),
            "sheets": [],
        }
    except Exception as exc:
        return {"status": "error", "message": str(exc), "sheets": []}

# ── AI Summary ─────────────────────────────────────────────
def get_ai_summary(df, quality_report):
    num_cols = df.select_dtypes(include="number").columns.tolist()
    cat_cols = df.select_dtypes(include="object").columns.tolist()
    date_cols = df.select_dtypes(include=["datetime64"]).columns.tolist()
    sample_stats = {}
    for col in num_cols[:5]:
        s = df[col].dropna()
        sample_stats[col] = {"sum": round(float(s.sum()),2), "mean": round(float(s.mean()),2),
                              "min": round(float(s.min()),2), "max": round(float(s.max()),2)}
    prompt = f"""You are a senior data analyst. Analyze this dataset and write a concise plain-English summary (5-7 sentences).

Dataset overview:
- Rows: {len(df)}, Columns: {len(df.columns)}
- Numeric columns: {num_cols}
- Categorical columns: {cat_cols}
- Date columns: {date_cols}
- Numeric stats: {json.dumps(sample_stats)}
- Quality issues found: {quality_report['issues'][:5]}
- Outlier columns: {list(quality_report['outliers'].keys())}

Write:
1. What this dataset appears to contain
2. Key numeric highlights (totals, averages)
3. Any data quality concerns
4. What analysis would be most valuable
Keep it practical and analyst-focused. No bullet points, just paragraphs."""
    try:
        resp = http_requests.post("https://api.anthropic.com/v1/messages",
            headers={"Content-Type": "application/json"},
            json={"model": "claude-sonnet-4-20250514", "max_tokens": 600,
                  "messages": [{"role": "user", "content": prompt}]},
            timeout=30)
        data = resp.json()
        return data["content"][0]["text"]
    except Exception as e:
        return f"AI summary unavailable: {e}"

# ════════════════════════════════════════════════════════════
#  UI START
# ════════════════════════════════════════════════════════════

st.markdown("""
<div class="dashboard-header">
  <div class="header-icon">📊</div>
  <div>
    <h1>Excel Analyst Dashboard</h1>
    <p>Upload · Profile · Clean · Query · Visualize · Export · Store — Powered by FastAPI + SQLite</p>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Sidebar ────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Settings")
    mask_pii_flag = st.toggle("🔒 Mask PII columns", value=False)
    show_ai = st.toggle("🤖 AI Data Summary", value=True)
    max_rows_preview = st.slider("Preview rows", 20, 500, 100, 10)

    st.markdown("---")
    st.markdown("### 🗄️ Load Saved Dataset")
    saved_tables = get_saved_tables()
    if saved_tables:
        load_table = st.selectbox("Select saved table", ["— select —"] + saved_tables)
        if load_table != "— select —":
            if st.button("📥 Load Table"):
                st.session_state.loaded_df = load_table_from_db(load_table)
                st.session_state.loaded_table_name = load_table
                st.success(f"Loaded `{load_table}`")
    else:
        st.caption("No saved tables yet.")

    st.markdown("---")
    st.markdown("### 📋 Upload History")
    if "upload_history" in st.session_state and st.session_state.upload_history:
        for h in st.session_state.upload_history[:5]:
            st.markdown(f"**{h['file']}** › {h['sheet']}<br><small>{h['rows']:,}R × {h['cols']}C &nbsp;|&nbsp; {h['time']}</small>", unsafe_allow_html=True)
    else:
        st.caption("No uploads yet.")

# ── File Upload ────────────────────────────────────────────
st.markdown('<div class="section-header"><span class="section-icon">📁</span> Upload Excel File</div>', unsafe_allow_html=True)
uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx","xls"], label_visibility="collapsed")

if "loaded_df" in st.session_state and uploaded_file is None:
    st.markdown(f'<div class="custom-info">📥 Showing loaded table: <b>{st.session_state.loaded_table_name}</b></div>', unsafe_allow_html=True)
    df_raw = st.session_state.loaded_df
    selected_sheet = st.session_state.loaded_table_name
    uploaded_file_name = selected_sheet
    process_data = True
elif uploaded_file:
    process_data = True
    uploaded_file_name = uploaded_file.name
else:
    process_data = False

if process_data:
    # ── Parse Sheet ──────────────────────────────────────
    if uploaded_file:
        xl = pd.ExcelFile(uploaded_file)
        sheet_names = xl.sheet_names
        st.markdown('<div class="section-header"><span class="section-icon">📋</span> Select Sheet</div>', unsafe_allow_html=True)
        if len(sheet_names) == 1:
            selected_sheet = sheet_names[0]
            st.markdown(f'<span class="badge">1 sheet: <b>{selected_sheet}</b></span>', unsafe_allow_html=True)
        else:
            selected_sheet = st.selectbox(f"📄 {len(sheet_names)} sheets found — select one to preview:", sheet_names)
        df_raw, hdr_row = detect_and_parse_sheet(xl, selected_sheet)
        if hdr_row > 0:
            st.markdown(f'<div class="custom-warning">⚠️ <b>{hdr_row} junk rows</b> skipped — real headers at row {hdr_row+1} ✅</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="custom-success">✅ Clean file — headers at row 1.</div>', unsafe_allow_html=True)
        log_upload(uploaded_file.name, selected_sheet, len(df_raw), len(df_raw.columns))

        # ════════════════════════════════════════════════════
        #  IMPORT SECTION
        #  Clicking "Import All Sheets" sends the file to FastAPI
        #  which runs the SAME clean_data() logic and persists
        #  every sheet as a separate SQLite table (REPLACE).
        # ════════════════════════════════════════════════════
        st.markdown("""
        <div class="import-btn-wrap">
          <div style="flex:1;">
            <h3>📥 Import All Sheets to Database</h3>
            <p>Parses, cleans (dedup · fill missing · normalise columns) and saves every tab as a separate SQLite table — identical to the old "Clean &amp; Save" but for <em>all</em> sheets at once.</p>
          </div>
        </div>
        """, unsafe_allow_html=True)

        import_col, status_col_ui = st.columns([1, 3])
        with import_col:
            do_import = st.button("⬆ Import All Sheets", use_container_width=True)

        if do_import:
            uploaded_file.seek(0)
            file_bytes = uploaded_file.read()
            with st.spinner(f"Cleaning & saving {len(sheet_names)} sheet(s) to database via FastAPI…"):
                result = call_fastapi_import(file_bytes, uploaded_file.name)

            if result.get("status") == "success":
                st.markdown(
                    f'<div class="custom-success">✅ <b>Import complete!</b> '
                    f'{result.get("sheets_imported", 0)} of {result.get("sheets_total", 0)} '
                    f'sheets cleaned and saved to the database.</div>',
                    unsafe_allow_html=True
                )
                for sheet_info in result.get("sheets", []):
                    ok = sheet_info.get("status") == "ok"
                    status_html = (
                        f'<span class="status-ok">✔ saved</span>'
                        if ok else
                        f'<span class="status-err">✘ {sheet_info.get("error","error")}</span>'
                    )
                    st.markdown(
                        f'<div class="import-result-row">'
                        f'  <span class="sheet-name">📄 {sheet_info["sheet"]}</span>'
                        f'  <span class="row-count">{sheet_info.get("rows",0):,} rows · {sheet_info.get("cols",0)} cols</span>'
                        f'  <span>→ table: <code>{sheet_info.get("table_name","")}</code></span>'
                        f'  {status_html}'
                        f'</div>',
                        unsafe_allow_html=True
                    )
                # Refresh sidebar table list
                st.rerun()
            else:
                st.markdown(
                    f'<div class="custom-error">❌ Import failed: {result.get("message","Unknown error")}</div>',
                    unsafe_allow_html=True
                )
                if "not running" in result.get("message",""):
                    st.code("uvicorn api:app --reload --port 8000", language="bash")

    # ── PII Detection ────────────────────────────────────
    pii_cols = detect_pii_columns(df_raw)
    if pii_cols:
        pii_str = ", ".join([f"<span class='pii-badge'>🔐 {c}</span>" for c in pii_cols])
        st.markdown(f'<div class="custom-warning">🔐 <b>PII columns detected:</b> {pii_str}</div>', unsafe_allow_html=True)
    if mask_pii_flag and pii_cols:
        df_raw = mask_pii(df_raw, pii_cols)
        st.markdown('<div class="custom-success">🔒 PII columns are masked.</div>', unsafe_allow_html=True)

    num_cols_list  = df_raw.select_dtypes(include="number").columns.tolist()
    cat_cols_list  = df_raw.select_dtypes(include="object").columns.tolist()
    date_cols_list = df_raw.select_dtypes(include=["datetime64"]).columns.tolist()

    # ════ TABS ═══════════════════════════════════════════
    tabs = st.tabs(["📋 Overview", "📈 Metrics", "🔬 Quality",
                    "📊 Charts", "🏗️ Group & Pivot", "🔍 Query",
                    "🤖 AI Summary", "🧹 Clean", "📤 Export", "📑 Reports"])

    # ── TAB 1: OVERVIEW ──────────────────────────────────
    with tabs[0]:
        missing_total = df_raw.isnull().sum().sum()
        missing_pct = missing_total / (df_raw.shape[0]*df_raw.shape[1])*100 if df_raw.size > 0 else 0
        c1,c2,c3,c4,c5,c6 = st.columns(6)
        for col_obj, lbl, val, sub in [
            (c1,"Total Rows",f"{len(df_raw):,}",""),
            (c2,"Columns",f"{len(df_raw.columns)}",""),
            (c3,"Numeric",f"{len(num_cols_list)}",""),
            (c4,"Categorical",f"{len(cat_cols_list)}",""),
            (c5,"Date Cols",f"{len(date_cols_list)}",""),
            (c6,"Missing",f"{int(missing_total):,}",f"{missing_pct:.1f}% of cells"),
        ]:
            with col_obj:
                sub_h = f'<div class="metric-detail">{sub}</div>' if sub else ""
                st.markdown(f'<div class="metric-card"><div class="metric-label">{lbl}</div><div class="metric-value">{val}</div>{sub_h}</div>', unsafe_allow_html=True)

        st.markdown('<div class="section-header"><span class="section-icon">📄</span> Column Profile</div>', unsafe_allow_html=True)
        col_info = pd.DataFrame({
            "Column":   df_raw.columns,
            "Type":     df_raw.dtypes.astype(str).values,
            "Non-Null": df_raw.notnull().sum().values,
            "Null %":   (df_raw.isnull().sum().values / len(df_raw) * 100).round(1),
            "Unique":   [df_raw[c].nunique() for c in df_raw.columns],
            "Sample":   [str(df_raw[c].dropna().iloc[0]) if df_raw[c].notnull().any() else "—" for c in df_raw.columns],
        })
        st.dataframe(col_info, use_container_width=True, hide_index=True)
        st.markdown('<div class="section-header"><span class="section-icon">📄</span> Raw Data Preview</div>', unsafe_allow_html=True)
        st.dataframe(df_raw.head(max_rows_preview), use_container_width=True)

    # ── TAB 2: METRICS ────────────────────────────────────
    with tabs[1]:
        pure_num_cols = [
            c for c in num_cols_list
            if c not in date_cols_list
            and not any(kw in c.lower() for kw in ["date","time","year","month","day","timestamp"])
        ]
        if not pure_num_cols:
            st.info("No numeric columns found.")
        else:
            sel_cols = st.multiselect("Select columns to analyse:", pure_num_cols,
                                       default=pure_num_cols[:min(4, len(pure_num_cols))])
            if sel_cols:
                st.markdown('<div class="section-header"><span class="section-icon">📈</span> Totals & Averages</div>', unsafe_allow_html=True)
                for i in range(0, len(sel_cols), 3):
                    row_cols = sel_cols[i:i+3]
                    ui_cols = st.columns(len(row_cols))
                    for j, col in enumerate(row_cols):
                        s = compute_metric_stats(df_raw[col])
                        with ui_cols[j]:
                            st.markdown(f"""
                            <div class="metric-card">
                              <div class="metric-label">{col}</div>
                              <div class="metric-value">{s['total']:,.2f}</div>
                              <div class="metric-sub sub-blue">Avg: {s['mean']:,.2f}</div>
                            </div>""", unsafe_allow_html=True)

                st.markdown('<div class="section-header"><span class="section-icon">📊</span> Percentage Contribution Table</div>', unsafe_allow_html=True)
                pct_rows = []
                grand_total = sum(df_raw[c].dropna().sum() for c in sel_cols)
                for col in sel_cols:
                    s = compute_metric_stats(df_raw[col])
                    col_total = s["total"]
                    pct_of_grand = (col_total / grand_total * 100) if grand_total != 0 else 0
                    pct_rows.append({
                        "Column": col, "Total": round(col_total, 2), "Average": round(s["mean"], 2),
                        "% of Grand Total": round(pct_of_grand, 2), "Non-null Rows": s["count"],
                        "Data Completeness %": round(s["non_null_pct"], 1),
                    })
                pct_df = pd.DataFrame(pct_rows)
                st.dataframe(pct_df, use_container_width=True, hide_index=True)
                fig_pct = px.bar(pct_df, x="Column", y="% of Grand Total",
                                 title="% Contribution of Each Column to Grand Total",
                                 text="% of Grand Total", color="Column",
                                 color_discrete_sequence=px.colors.qualitative.Pastel)
                fig_pct.update_traces(texttemplate="%{text:.1f}%", textposition="outside")
                fig_pct.update_layout(plot_bgcolor="#fff", paper_bgcolor="#fff",
                                      font_family="DM Sans", showlegend=False)
                st.plotly_chart(fig_pct, use_container_width=True)

    # ── TAB 3: QUALITY ────────────────────────────────────
    with tabs[2]:
        qr = data_quality_report(df_raw)
        if qr["issues"]:
            st.markdown('<div class="section-header"><span class="section-icon">⚠️</span> Issues Found</div>', unsafe_allow_html=True)
            for issue in qr["issues"]:
                st.markdown(f'<div class="custom-warning">{issue}</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="custom-success">✅ No data quality issues detected.</div>', unsafe_allow_html=True)
        if qr["outliers"]:
            st.markdown('<div class="section-header"><span class="section-icon">🎯</span> Outlier Details (IQR Method)</div>', unsafe_allow_html=True)
            for col, info in qr["outliers"].items():
                with st.expander(f"📌 {col} — {info['count']} outliers ({info['pct']:.1f}%)"):
                    out_df = df_raw.iloc[info["rows"]]
                    st.dataframe(out_df, use_container_width=True)
                    fig = px.box(df_raw, y=col, title=f"Box Plot — {col}",
                                 color_discrete_sequence=["#6366f1"])
                    fig.update_layout(plot_bgcolor="#fff", paper_bgcolor="#fff", font_family="DM Sans")
                    st.plotly_chart(fig, use_container_width=True)

    # ── TAB 4: CHARTS ─────────────────────────────────────
    with tabs[3]:
        chart_tabs = st.tabs(["🤖 Auto Charts", "🏗️ Chart Builder"])
        with chart_tabs[0]:
            charts = auto_charts(df_raw)
            if charts:
                chart_names = [c[0] for c in charts]
                sel = st.selectbox("Select chart:", chart_names)
                for title, fig in charts:
                    if title == sel:
                        st.plotly_chart(fig, use_container_width=True)
                        break
            else:
                st.warning("No charts available — check data types.")
        with chart_tabs[1]:
            all_cols = df_raw.columns.tolist()
            cb1, cb2, cb3, cb4 = st.columns(4)
            with cb1:
                chart_type = st.selectbox("Chart type", ["Bar","Line","Scatter","Histogram","Box","Pie","Area"])
            with cb2:
                x_col = st.selectbox("X axis", all_cols)
            with cb3:
                y_col = st.selectbox("Y axis", ["— none —"] + num_cols_list)
            with cb4:
                color_col = st.selectbox("Color by", ["— none —"] + cat_cols_list)
            color_arg = None if color_col == "— none —" else color_col
            y_arg = None if y_col == "— none —" else y_col
            agg_type = st.selectbox("Aggregation", ["none","sum","mean","count","min","max"])
            if st.button("📊 Generate Chart"):
                try:
                    plot_df = df_raw.copy()
                    if agg_type != "none" and y_arg and chart_type in ("Bar","Line","Area"):
                        if agg_type == "count":
                            plot_df = plot_df.groupby(x_col).size().reset_index(name="count")
                            y_arg = "count"
                        else:
                            plot_df = plot_df.groupby(x_col, as_index=False)[y_arg].agg(agg_type)
                    layout = dict(plot_bgcolor="#fff", paper_bgcolor="#fff", font_family="DM Sans")
                    fig_map = {
                        "Bar":       px.bar(plot_df, x=x_col, y=y_arg, color=color_arg),
                        "Line":      px.line(plot_df.sort_values(x_col), x=x_col, y=y_arg, color=color_arg),
                        "Area":      px.area(plot_df.sort_values(x_col), x=x_col, y=y_arg, color=color_arg),
                        "Scatter":   px.scatter(plot_df, x=x_col, y=y_arg, color=color_arg),
                        "Histogram": px.histogram(plot_df, x=x_col, color=color_arg, nbins=30),
                        "Box":       px.box(plot_df, x=color_arg, y=y_arg or x_col),
                        "Pie":       px.pie(plot_df[x_col].value_counts().reset_index().rename(columns={x_col:"val","count":"cnt"}), names="val", values="cnt"),
                    }
                    fig = fig_map.get(chart_type, fig_map["Bar"])
                    fig.update_layout(**layout)
                    st.plotly_chart(fig, use_container_width=True)
                except Exception as e:
                    st.error(f"Chart error: {e}")

    # ── TAB 5: GROUP & PIVOT ──────────────────────────────
    with tabs[4]:
        gb_tabs = st.tabs(["📊 Group By", "🔄 Pivot Table"])
        with gb_tabs[0]:
            g1, g2, g3 = st.columns(3)
            with g1:
                gb_group_cols = st.multiselect("Group by:", cat_cols_list + date_cols_list,
                                                default=cat_cols_list[:1] if cat_cols_list else [])
            with g2:
                gb_agg_col = st.selectbox("Aggregate:", num_cols_list) if num_cols_list else None
            with g3:
                gb_func = st.selectbox("Function:", ["sum","mean","count","min","max","median","std"])
            if st.button("▶ Run Group By") and gb_group_cols and gb_agg_col:
                try:
                    if gb_func == "count":
                        gb_result = df_raw.groupby(gb_group_cols).size().reset_index(name="count")
                    else:
                        gb_result = df_raw.groupby(gb_group_cols, as_index=False)[gb_agg_col].agg(gb_func)
                    gb_result = gb_result.sort_values(gb_result.columns[-1], ascending=False)
                    st.dataframe(gb_result, use_container_width=True, hide_index=True)
                    if len(gb_group_cols) == 1 and len(gb_result) <= 30:
                        fig = px.bar(gb_result, x=gb_group_cols[0], y=gb_result.columns[-1],
                                     color_discrete_sequence=["#6366f1"])
                        fig.update_layout(plot_bgcolor="#fff", paper_bgcolor="#fff", font_family="DM Sans")
                        st.plotly_chart(fig, use_container_width=True)
                    st.download_button("⬇️ Download CSV", gb_result.to_csv(index=False).encode(),
                                       "groupby_result.csv", mime="text/csv")
                except Exception as e:
                    st.error(f"Group By error: {e}")
        with gb_tabs[1]:
            p1, p2, p3, p4 = st.columns(4)
            with p1:
                pvt_rows = st.multiselect("Row(s):", cat_cols_list, default=cat_cols_list[:1] if cat_cols_list else [])
            with p2:
                pvt_cols = st.multiselect("Column(s):", cat_cols_list, default=[])
            with p3:
                pvt_val = st.selectbox("Values:", num_cols_list) if num_cols_list else None
            with p4:
                pvt_func = st.selectbox("Aggregation:", ["sum","mean","count","min","max"])
            if st.button("▶ Build Pivot") and pvt_rows and pvt_val:
                try:
                    pivot = pd.pivot_table(df_raw, values=pvt_val, index=pvt_rows,
                                           columns=pvt_cols if pvt_cols else None,
                                           aggfunc=pvt_func, fill_value=0,
                                           margins=True, margins_name="TOTAL")
                    st.dataframe(pivot.round(2), use_container_width=True)
                    st.download_button("⬇️ Download Pivot", pivot.to_csv().encode(), "pivot.csv", mime="text/csv")
                except Exception as e:
                    st.error(f"Pivot error: {e}")

    # ── TAB 6: QUERY ─────────────────────────────────────
    with tabs[5]:
        col_names_str = ", ".join([f"`{c}`" for c in df_raw.columns.tolist()])
        if "query_history" not in st.session_state:
            st.session_state.query_history = []

        with st.expander("📖 Query Reference"):
            st.markdown(f"""
**Available Columns:** {col_names_str}

| Pattern | Example |
|---|---|
| Filter | `Status == 'Active'` |
| Aggregation | `SUM(Cost)` |
| Agg + Filter | `SUM(Cost) WHERE Status == 'Active'` |
| Group By | `COUNT(*) GROUP BY Status` |
| Top N | `TOP 10 BY Cost DESC` |
""")
        q_input = st.text_input("Query:", placeholder="e.g. TOP 10 BY Cost  or  Status == 'Active'", key="query_input")
        run_col, hist_col = st.columns([1,3])
        with run_col:
            run_q = st.button("▶ Run Query", use_container_width=True)
        with hist_col:
            if st.session_state.query_history:
                hist_sel = st.selectbox("📜 Recent:", ["— select —"] + st.session_state.query_history[:15])
                if hist_sel != "— select —":
                    q_input = hist_sel

        if run_q and q_input.strip():
            q_clean = q_input.strip()
            st.session_state.query_history = ([q_clean] + [x for x in st.session_state.query_history if x != q_clean])[:15]
            result_df, msg = run_custom_query(df_raw, q_clean)
            if result_df is not None:
                st.markdown(f'<div class="custom-success">{msg}</div>', unsafe_allow_html=True)
                st.dataframe(result_df, use_container_width=True, hide_index=True)
                dl1, dl2 = st.columns(2)
                with dl1:
                    st.download_button("⬇️ CSV", result_df.to_csv(index=False).encode(),
                                       "query_result.csv", mime="text/csv", use_container_width=True)
                with dl2:
                    xl_buf = export_to_excel(result_df, "Query Result")
                    st.download_button("⬇️ Excel", xl_buf, "query_result.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True)
            else:
                st.markdown(f'<div class="custom-error">{msg}</div>', unsafe_allow_html=True)

    # ── TAB 7: AI SUMMARY ────────────────────────────────
    with tabs[6]:
        if show_ai:
            if st.button("🤖 Generate AI Summary"):
                with st.spinner("Analyzing with Claude AI..."):
                    qr_for_ai = data_quality_report(df_raw)
                    summary = get_ai_summary(df_raw, qr_for_ai)
                st.markdown(f'<div class="ai-box"><h4>🤖 Claude AI Analysis</h4><p>{summary}</p></div>', unsafe_allow_html=True)

            nl_query = st.text_input("Ask anything about your data:", placeholder="e.g. What are the top 5 rows by cost?")
            if st.button("💬 Ask AI") and nl_query.strip():
                with st.spinner("Thinking..."):
                    col_info_str = ", ".join([f"{c} ({str(df_raw[c].dtype)})" for c in df_raw.columns])
                    sample = df_raw.head(3).to_dict(orient="records")
                    prompt = f"""You are a data analyst. The user has a DataFrame with columns: {col_info_str}
Sample rows: {json.dumps(sample, default=str)}
User question: {nl_query}
Answer concisely. If the question needs a calculation, provide the pandas code AND the answer."""
                    try:
                        resp = http_requests.post("https://api.anthropic.com/v1/messages",
                            headers={"Content-Type": "application/json"},
                            json={"model": "claude-sonnet-4-20250514", "max_tokens": 500,
                                  "messages": [{"role": "user", "content": prompt}]},
                            timeout=30)
                        answer = resp.json()["content"][0]["text"]
                        st.markdown(f'<div class="ai-box"><h4>💬 Answer</h4><p>{answer}</p></div>', unsafe_allow_html=True)
                    except Exception as e:
                        st.error(f"AI error: {e}")
        else:
            st.info("Enable AI Summary in the sidebar settings.")

    # ── TAB 8: CLEAN (preview only — DB saving via Import button) ──
    with tabs[7]:
        st.markdown(
            '<div class="custom-info">ℹ️ Click <b>🧹 Clean Data</b> to preview the cleaned result. '
            'To persist to the database, use the <b>⬆ Import All Sheets</b> button above — '
            'it runs the same cleaning logic and saves every sheet to SQLite automatically.</div>',
            unsafe_allow_html=True
        )

        if st.button("🧹 Clean Data"):
            with st.spinner("Cleaning data…"):
                df_clean, profile = clean_data(df_raw.copy())
            # Store in session so the preview survives a rerun
            st.session_state["last_cleaned_df"]      = df_clean
            st.session_state["last_cleaned_profile"] = profile

        if "last_cleaned_df" in st.session_state:
            df_clean = st.session_state["last_cleaned_df"]
            profile  = st.session_state["last_cleaned_profile"]

            st.markdown(
                '<div class="custom-success">✅ Data cleaned — preview below. '
                'Use <b>⬆ Import All Sheets</b> to save to the database.</div>',
                unsafe_allow_html=True
            )
            s1, s2, s3, s4 = st.columns(4)
            for col_obj, lbl, val in [
                (s1, "Original Rows",    f"{profile['Original Rows']:,}"),
                (s2, "Cleaned Rows",     f"{profile['Cleaned Rows']:,}"),
                (s3, "Dupes Removed",    f"{profile['Duplicates Removed']:,}"),
                (s4, "Missing After",    f"{profile['Missing After']:,}"),
            ]:
                with col_obj:
                    st.markdown(
                        f'<div class="metric-card">'
                        f'<div class="metric-label">{lbl}</div>'
                        f'<div class="metric-value" style="font-size:1.5rem">{val}</div>'
                        f'</div>',
                        unsafe_allow_html=True
                    )
            st.dataframe(df_clean.head(100), use_container_width=True)
            if st.button("💾 Save to MySQL (port 3307)"):
                try:
                    saved_name = save_df_to_mysql(df_clean, selected_sheet)
                    st.markdown(
                        f'<div class="custom-success">✅ Saved to MySQL port 3307 as table: <b>{saved_name}</b></div>',
                        unsafe_allow_html=True
                    )
                    st.rerun()
                except Exception as e:
                    st.markdown(
                        f'<div class="custom-error">❌ MySQL save failed: {e}</div>',
                        unsafe_allow_html=True
                    )


    # ── TAB 9: EXPORT ────────────────────────────────────
    with tabs[8]:
        e1, e2, e3 = st.columns(3)
        with e1:
            st.markdown("**📄 CSV**")
            st.download_button("⬇️ Download CSV", df_raw.to_csv(index=False).encode(),
                               "data.csv", mime="text/csv", use_container_width=True)
        with e2:
            st.markdown("**📊 Excel (Formatted)**")
            xl_buf = export_to_excel(df_raw, selected_sheet)
            st.download_button("⬇️ Download Excel", xl_buf, "data_export.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        with e3:
            st.markdown("**📋 JSON**")
            json_str = df_raw.to_json(orient="records", date_format="iso", indent=2)
            st.download_button("⬇️ Download JSON", json_str.encode(),
                               "data.json", mime="application/json", use_container_width=True)
        if num_cols_list:
            st.markdown('<div class="section-header"><span class="section-icon">📐</span> Summary Stats</div>', unsafe_allow_html=True)
            stats_df = df_raw[num_cols_list].describe().T.round(3)
            st.dataframe(stats_df, use_container_width=True)

    # ── TAB 10: REPORTS ──────────────────────────────────
    with tabs[9]:
        st.markdown('<div class="section-header"><span class="section-icon">📑</span> Cross-Sheet Reports</div>', unsafe_allow_html=True)
        if uploaded_file:
            all_sheets = {}
            try:
                xl_all = pd.ExcelFile(uploaded_file)
                for sht in xl_all.sheet_names:
                    df_sht, _ = detect_and_parse_sheet(xl_all, sht)
                    all_sheets[sht] = df_sht
            except Exception as e:
                st.error(f"Could not load all sheets: {e}")

            def find_col(df, keywords):
                for kw in keywords:
                    for col in df.columns:
                        if kw.lower() in col.lower():
                            return col
                return None

            def sheet_for(keywords):
                for kw in keywords:
                    for sht in all_sheets:
                        if kw.lower() in sht.lower():
                            return sht, all_sheets[sht]
                return None, None

            proj_sht, df_proj = sheet_for(["project","overview","proj"])
            task_sht, df_task = sheet_for(["task","tasks","work"])
            res_sht,  df_res  = sheet_for(["resource","employee","staff","team"])
            cr_sht,   df_cr   = sheet_for(["change","cr","request"])

            if df_proj is None and len(all_sheets) == 1:
                only_sht = list(all_sheets.keys())[0]
                df_proj = all_sheets[only_sht]

            r1, r2, r3 = st.tabs(["📋 Task Count", "👥 Resource Count", "🔄 Change Requests"])

            with r1:
                df_t = df_task if df_task is not None else df_proj
                if df_t is not None:
                    proj_id_col = find_col(df_t, ["project_id","project id","proj_id","project"])
                    if proj_id_col:
                        task_count = df_t.groupby(proj_id_col).size().reset_index(name="Task Count")
                        task_count = task_count.sort_values("Task Count", ascending=False)
                        st.dataframe(task_count, use_container_width=True)
                        fig = px.bar(task_count, x=proj_id_col, y="Task Count",
                                     title="Task Count per Project", color="Task Count",
                                     color_continuous_scale="Blues", text="Task Count")
                        fig.update_layout(plot_bgcolor="#fff", paper_bgcolor="#fff", font_family="DM Sans")
                        st.plotly_chart(fig, use_container_width=True)
                        st.download_button("⬇️ CSV", task_count.to_csv(index=False).encode(),
                                           "report1_task_count.csv", mime="text/csv")
                    else:
                        st.warning(f"No Project ID column found. Available: {list(df_t.columns)}")
                else:
                    st.warning("No task/project sheet found.")

            with r2:
                df_r = df_res if df_res is not None else df_proj
                if df_r is not None:
                    proj_id_col = find_col(df_r, ["project_id","project id","proj_id","project"])
                    if proj_id_col:
                        res_count = df_r.groupby(proj_id_col).size().reset_index(name="Resource Count")
                        res_count = res_count.sort_values("Resource Count", ascending=False)
                        st.dataframe(res_count, use_container_width=True)
                        fig = px.bar(res_count, x=proj_id_col, y="Resource Count",
                                     title="Resource Count per Project", color="Resource Count",
                                     color_continuous_scale="Greens", text="Resource Count")
                        fig.update_layout(plot_bgcolor="#fff", paper_bgcolor="#fff", font_family="DM Sans")
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.warning(f"No Project ID column found. Available: {list(df_r.columns)}")
                else:
                    st.warning("No resource sheet found.")

            with r3:
                df_c = df_cr if df_cr is not None else df_proj
                if df_c is not None:
                    proj_id_col = find_col(df_c, ["project_id","project id","proj_id","project"])
                    if proj_id_col:
                        cr_count = df_c.groupby(proj_id_col).size().reset_index(name="Change Request Count")
                        cr_count = cr_count.sort_values("Change Request Count", ascending=False)
                        st.dataframe(cr_count, use_container_width=True)
                        fig = px.bar(cr_count, x=proj_id_col, y="Change Request Count",
                                     title="Change Requests per Project", color="Change Request Count",
                                     color_continuous_scale="Oranges", text="Change Request Count")
                        fig.update_layout(plot_bgcolor="#fff", paper_bgcolor="#fff", font_family="DM Sans")
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.warning(f"No Project ID column found. Available: {list(df_c.columns)}")
                else:
                    st.warning("No change request sheet found.")
        else:
            st.info("Upload a file to generate cross-sheet reports.")

else:
    st.markdown("""
    <div style="text-align:center;padding:80px 20px;color:#9ca3af;">
        <div style="font-size:3.5rem;margin-bottom:14px;">📂</div>
        <div style="font-size:1.2rem;font-weight:700;color:#6b7280;">Upload an Excel file to get started</div>
        <div style="font-size:.9rem;margin-top:8px;">or load a saved table from the sidebar</div>
        <div style="font-size:.82rem;margin-top:4px;color:#d1d5db;">Supports .xlsx and .xls</div>
    </div>
    """, unsafe_allow_html=True)
